import openpyxl

def getRowCount(filepath, sheet_no):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheet_no]
    return sheet.max_row

def getColumnCount(filepath, sheet_no):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheet_no]
    return sheet.max_column

def readData(filepath, sheet_no, row, col):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheet_no]
    return sheet.cell(row=row, column=col).value

def writeData(filepath, sheet_no, row, col, data):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook[sheet_no]
    sheet.cell(row=row, column=col).value = data
    workbook.save(filepath)